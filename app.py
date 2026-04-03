from flask import Flask, render_template, request, redirect, session, url_for
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
import os

app = Flask(__name__)
app.secret_key = "secret123"

app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///database.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
db = SQLAlchemy(app)


class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    exam_timer = db.Column(db.Integer, nullable=False, default=120)


class Question(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    subject = db.Column(db.String(100), nullable=False)
    question = db.Column(db.String(500), nullable=False)
    opt1 = db.Column(db.String(200), nullable=False)
    opt2 = db.Column(db.String(200), nullable=False)
    opt3 = db.Column(db.String(200), nullable=False)
    opt4 = db.Column(db.String(200), nullable=False)
    answer = db.Column(db.String(200), nullable=False)
    explanation = db.Column(db.String(1000), nullable=True)


class Result(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    student_id = db.Column(db.String(100), nullable=False)
    subject = db.Column(db.String(100), nullable=False)
    score = db.Column(db.Integer, nullable=False)
    total = db.Column(db.Integer, nullable=False)


class AdminAccount(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(100), nullable=False, default="admin")
    password = db.Column(db.String(100), nullable=False, default="admin123")


def calculate_grade(percentage: float) -> str:
    if percentage >= 70:
        return "A"
    if percentage >= 60:
        return "B"
    if percentage >= 50:
        return "C"
    if percentage >= 45:
        return "D"
    if percentage >= 40:
        return "E"
    return "F"


@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        student_name = request.form.get("name", "").strip()
        student_id = request.form.get("student_id", "").strip()
        admin_username = request.form.get("admin_username", "").strip()
        admin_password = request.form.get("admin_password", "").strip()

        admin_account = AdminAccount.query.first()

        if admin_account and admin_username == admin_account.username and admin_password == admin_account.password:
            session.clear()
            session["admin"] = True
            return redirect(url_for("admin"))

        if student_name and student_id:
            session.clear()
            session["student"] = student_name
            session["student_id"] = student_id
            return redirect(url_for("home"))

    return render_template("login.html")


@app.route("/home")
def home():
    if "student" not in session:
        return redirect(url_for("login"))

    subjects = Subject.query.order_by(Subject.name.asc()).all()
    return render_template(
        "home.html",
        subjects=subjects,
        student_name=session["student"]
    )


@app.route("/student/<int:subject_id>")
def student(subject_id):
    if "student" not in session:
        return redirect(url_for("login"))

    subject = Subject.query.get_or_404(subject_id)
    questions = Question.query.filter_by(subject=subject.name).order_by(Question.id.asc()).all()

    if not questions:
        return redirect(url_for("home"))

    session["current_subject"] = subject.name

    return render_template(
        "student.html",
        questions=questions,
        exam_timer=subject.exam_timer,
        student_name=session["student"],
        subject_name=subject.name
    )


@app.route("/submit", methods=["POST"])
def submit():
    if "student" not in session:
        return redirect(url_for("login"))

    subject_name = session.get("current_subject")
    if not subject_name:
        return redirect(url_for("home"))

    questions = Question.query.filter_by(subject=subject_name).order_by(Question.id.asc()).all()
    if not questions:
        return redirect(url_for("home"))

    score = 0
    total = len(questions)
    review_rows = []

    for q in questions:
        selected = request.form.get(str(q.id), "").strip()
        correct = (q.answer or "").strip()
        is_correct = selected and correct and selected == correct

        if is_correct:
            score += 1

        review_rows.append({
            "question": q.question,
            "your_answer": selected if selected else "Not Answered",
            "correct_answer": correct,
            "explanation": q.explanation if q.explanation else "No explanation provided.",
            "is_correct": bool(is_correct)
        })

    percentage = round((score / total) * 100, 2) if total else 0
    grade = calculate_grade(percentage)

    result = Result(
        name=session["student"],
        student_id=session["student_id"],
        subject=subject_name,
        score=score,
        total=total
    )
    db.session.add(result)
    db.session.commit()

    return render_template(
        "result.html",
        score=score,
        total=total,
        percentage=percentage,
        grade=grade,
        student_name=session["student"],
        subject_name=subject_name,
        review_rows=review_rows
    )


@app.route("/history")
def history():
    if "student" not in session:
        return redirect(url_for("login"))

    results = Result.query.filter_by(student_id=session["student_id"]).order_by(Result.id.desc()).all()

    rows = []
    for r in results:
        percentage = round((r.score / r.total) * 100, 2) if r.total else 0
        rows.append({
            "subject": r.subject,
            "score": r.score,
            "total": r.total,
            "percentage": percentage,
            "grade": calculate_grade(percentage)
        })

    return render_template(
        "history.html",
        student_name=session["student"],
        student_id=session["student_id"],
        results=rows
    )


@app.route("/leaderboard")
def leaderboard():
    results = Result.query.order_by(Result.score.desc(), Result.id.asc()).all()

    rows = []
    for r in results:
        percentage = round((r.score / r.total) * 100, 2) if r.total else 0
        rows.append({
            "name": r.name,
            "student_id": r.student_id,
            "subject": r.subject,
            "score": r.score,
            "total": r.total,
            "percentage": percentage,
            "grade": calculate_grade(percentage)
        })

    return render_template("leaderboard.html", results=rows)


@app.route("/admin")
def admin():
    if "admin" not in session:
        return redirect(url_for("login"))

    subjects = Subject.query.order_by(Subject.name.asc()).all()
    questions = Question.query.order_by(Question.id.desc()).all()
    admin_account = AdminAccount.query.first()

    return render_template(
        "admin.html",
        subjects=subjects,
        questions=questions,
        admin_account=admin_account
    )


@app.route("/update_admin_account", methods=["POST"])
def update_admin_account():
    if "admin" not in session:
        return redirect(url_for("login"))

    new_username = request.form.get("new_username", "").strip()
    new_password = request.form.get("new_password", "").strip()

    admin_account = AdminAccount.query.first()

    if admin_account and new_username and new_password:
        admin_account.username = new_username
        admin_account.password = new_password
        db.session.commit()

    return redirect(url_for("admin"))


@app.route("/add_subject", methods=["POST"])
def add_subject():
    if "admin" not in session:
        return redirect(url_for("login"))

    name = request.form.get("name", "").strip()
    exam_timer = request.form.get("exam_timer", "").strip()

    if name and exam_timer.isdigit():
        existing = Subject.query.filter_by(name=name).first()
        if not existing:
            db.session.add(
                Subject(
                    name=name,
                    exam_timer=int(exam_timer)
                )
            )
            db.session.commit()

    return redirect(url_for("admin"))


@app.route("/add", methods=["GET", "POST"])
def add():
    if "admin" not in session:
        return redirect(url_for("login"))

    subjects = Subject.query.order_by(Subject.name.asc()).all()

    if request.method == "POST":
        subject = request.form.get("subject", "").strip()
        question = request.form.get("question", "").strip()
        opt1 = request.form.get("opt1", "").strip()
        opt2 = request.form.get("opt2", "").strip()
        opt3 = request.form.get("opt3", "").strip()
        opt4 = request.form.get("opt4", "").strip()
        answer = request.form.get("answer", "").strip()
        explanation = request.form.get("explanation", "").strip()

        if subject and question and opt1 and opt2 and opt3 and opt4 and answer:
            db.session.add(
                Question(
                    subject=subject,
                    question=question,
                    opt1=opt1,
                    opt2=opt2,
                    opt3=opt3,
                    opt4=opt4,
                    answer=answer,
                    explanation=explanation
                )
            )
            db.session.commit()
            return redirect(url_for("admin"))

    return render_template("add.html", subjects=subjects)


@app.route("/import_excel", methods=["GET", "POST"])
def import_excel():
    if "admin" not in session:
        return redirect(url_for("login"))

    subjects = Subject.query.order_by(Subject.name.asc()).all()

    if request.method == "POST":
        excel_file = request.files.get("excel_file")
        selected_subject = request.form.get("subject", "").strip()

        if excel_file and selected_subject:
            upload_path = os.path.join("/tmp", excel_file.filename)
            excel_file.save(upload_path)

            workbook = load_workbook(upload_path)
            sheet = workbook.active

            # A=question, B=opt1, C=opt2, D=opt3, E=opt4, F=answer, G=explanation
            for row in sheet.iter_rows(min_row=2, values_only=True):
                question_text = row[0] if len(row) > 0 else None
                opt1 = row[1] if len(row) > 1 else None
                opt2 = row[2] if len(row) > 2 else None
                opt3 = row[3] if len(row) > 3 else None
                opt4 = row[4] if len(row) > 4 else None
                answer = row[5] if len(row) > 5 else None
                explanation = row[6] if len(row) > 6 else ""

                if question_text and opt1 and opt2 and opt3 and opt4 and answer:
                    db.session.add(
                        Question(
                            subject=selected_subject,
                            question=str(question_text).strip(),
                            opt1=str(opt1).strip(),
                            opt2=str(opt2).strip(),
                            opt3=str(opt3).strip(),
                            opt4=str(opt4).strip(),
                            answer=str(answer).strip(),
                            explanation=str(explanation).strip() if explanation else ""
                        )
                    )

            db.session.commit()
            return redirect(url_for("admin"))

    return render_template("import_excel.html", subjects=subjects)


@app.route("/delete/<int:id>")
def delete(id):
    if "admin" not in session:
        return redirect(url_for("login"))

    q = Question.query.get_or_404(id)
    db.session.delete(q)
    db.session.commit()
    return redirect(url_for("admin"))


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


with app.app_context():
    db.create_all()

    if AdminAccount.query.first() is None:
        db.session.add(AdminAccount(username="admin", password="admin123"))
        db.session.commit()


if __name__ == "__main__":
    app.run(debug=True)
